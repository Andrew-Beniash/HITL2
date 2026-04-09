"""
converter.py — PDF/Markdown-to-EPUB3 conversion using pdfminer.six + ebooklib.

Aligns with:
  - Functional Spec §5.1.1 (PDF display format)
  - Technical Architecture §3.2 (backend libraries)
  - Functional Spec §5.1.4 (client default font configuration defaults)
"""

from __future__ import annotations

import datetime
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from ebooklib import epub
from pdfminer.high_level import extract_pages
from pdfminer.layout import (
    LAParams,
    LTAnno,
    LTChar,
    LTFigure,
    LTPage,
    LTTextBox,
    LTTextLine,
)

# ── Platform default typography (Functional Spec §5.1.4) ──────────────────────
EPUB_STYLESHEET = """
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&family=JetBrains+Mono&display=swap');

body {
  font-family: 'Inter', system-ui, -apple-system, sans-serif;
  font-size: 1rem;
  line-height: 1.6;
  color: #1a1a1a;
  margin: 0 auto;
  max-width: 780px;
  padding: 2rem 1.5rem;
}

h1 { font-size: 2rem;   font-weight: 700; margin: 1.5rem 0 0.75rem; }
h2 { font-size: 1.5rem; font-weight: 700; margin: 1.25rem 0 0.6rem; }
h3 { font-size: 1.25rem;font-weight: 600; margin: 1rem 0 0.5rem; }

p  { margin: 0.6rem 0; }

pre, code {
  font-family: 'JetBrains Mono', monospace;
  font-size: 0.9rem;
  background: #f4f4f5;
  padding: 0.2em 0.4em;
  border-radius: 3px;
}

pre {
  padding: 1rem;
  overflow-x: auto;
  white-space: pre-wrap;
}

hr.page-break {
  border: none;
  border-top: 1px solid #e5e7eb;
  margin: 2rem 0;
}

.page-label {
  font-size: 0.75rem;
  color: #9ca3af;
  text-transform: uppercase;
  letter-spacing: 0.05em;
  margin-bottom: 0.5rem;
}

.table-wrap { overflow-x: auto; margin: 1.5rem 0; }
table { border-collapse: collapse; width: 100%; font-size: 0.875rem; }
th, td { border: 1px solid #e5e7eb; padding: 0.45rem 0.75rem; text-align: left; vertical-align: top; word-break: break-word; }
th { background: #f3f4f6; font-weight: 600; }
tr:nth-child(even) td { background: #f9fafb; }
.sheet-meta { font-size: 0.8rem; color: #6e6e73; margin-bottom: 1rem; }
"""

PAGES_PER_CHAPTER = 15  # group pages into chapters for long documents
MAX_FILE_SIZE_MB = 50


# ── Data classes ──────────────────────────────────────────────────────────────

@dataclass
class TextBlock:
    text: str
    font_size: float = 12.0
    is_bold: bool = False
    x0: float = 0.0
    y0: float = 0.0


@dataclass
class PageContent:
    number: int
    blocks: list[TextBlock] = field(default_factory=list)

    @property
    def is_empty(self) -> bool:
        return not any(b.text.strip() for b in self.blocks)


# ── PDF extraction ─────────────────────────────────────────────────────────────

def _char_is_bold(char: LTChar) -> bool:
    """Heuristic: font name contains Bold."""
    fn = getattr(char, "fontname", "") or ""
    return "bold" in fn.lower()


def _extract_pages_content(pdf_bytes: bytes) -> list[PageContent]:
    """Extract text blocks from all PDF pages using pdfminer."""
    laparams = LAParams(
        line_overlap=0.5,
        char_margin=2.0,
        line_margin=0.5,
        word_margin=0.1,
        boxes_flow=0.5,
        detect_vertical=False,
    )
    pages: list[PageContent] = []

    pdf_stream = io.BytesIO(pdf_bytes)
    for page_layout in extract_pages(pdf_stream, laparams=laparams):
        page_num = len(pages) + 1
        pc = PageContent(number=page_num)

        # Sort elements top-to-bottom, left-to-right
        elements = sorted(page_layout, key=lambda e: (-e.y1, e.x0))

        for element in elements:
            if not isinstance(element, LTTextBox):
                continue

            raw_text = element.get_text().strip()
            if not raw_text:
                continue

            # Sample font characteristics from first non-anon char
            sample_size = 12.0
            sample_bold = False
            for line in element:
                if not isinstance(line, LTTextLine):
                    continue
                for char in line:
                    if isinstance(char, LTChar):
                        sample_size = char.size
                        sample_bold = _char_is_bold(char)
                        break
                break

            pc.blocks.append(
                TextBlock(
                    text=raw_text,
                    font_size=round(sample_size, 1),
                    is_bold=sample_bold,
                    x0=element.x0,
                    y0=element.y0,
                )
            )

        pages.append(pc)

    return pages


# ── HTML generation ────────────────────────────────────────────────────────────

def _classify_block(block: TextBlock, body_size: float) -> str:
    """Return 'h1', 'h2', 'h3', or 'p' based on font size and bold flag."""
    ratio = block.font_size / body_size if body_size > 0 else 1.0
    if ratio >= 1.8 or (ratio >= 1.4 and block.is_bold):
        return "h2"
    if ratio >= 1.3 or (ratio >= 1.1 and block.is_bold):
        return "h3"
    return "p"


def _pages_to_html(pages: list[PageContent], chapter_title: str) -> str:
    """Convert a list of PageContent objects to an EPUB chapter HTML string."""
    if not pages:
        return f"<h2>{_esc(chapter_title)}</h2><p>(No text content on these pages.)</p>"

    # Estimate body font size as the median font size across all blocks
    all_sizes = [b.font_size for p in pages for b in p.blocks if b.text.strip()]
    if all_sizes:
        all_sizes.sort()
        body_size = all_sizes[len(all_sizes) // 2]
    else:
        body_size = 12.0

    parts: list[str] = [f"<h1>{_esc(chapter_title)}</h1>"]

    for i, page in enumerate(pages):
        if page.is_empty:
            continue

        if i > 0:
            parts.append(f'<hr class="page-break"/>')
            parts.append(f'<p class="page-label">Page {page.number}</p>')

        for block in page.blocks:
            text = block.text.strip()
            if not text:
                continue
            tag = _classify_block(block, body_size)
            # Split block into paragraphs on double newlines
            paragraphs = re.split(r"\n{2,}", text)
            for para in paragraphs:
                para = para.replace("\n", " ").strip()
                if para:
                    parts.append(f"<{tag}>{_esc(para)}</{tag}>")

    return "\n".join(parts)


def _esc(text: str) -> str:
    """Minimal HTML escaping."""
    return (
        text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
    )


# ── EPUB assembly ──────────────────────────────────────────────────────────────

def _make_book(stem: str) -> tuple[epub.EpubBook, epub.EpubItem]:
    book = epub.EpubBook()
    book.set_identifier(f"hitl-preview-{stem}")
    book.set_title(stem.replace("-", " ").replace("_", " ").title())
    book.set_language("en")
    book.add_author("HITL Preview")

    css = epub.EpubItem(
        uid="style_main",
        file_name="style/main.css",
        media_type="text/css",
        content=EPUB_STYLESHEET.encode("utf-8"),
    )
    book.add_item(css)
    return book, css

def _group_into_chapters(
    pages: list[PageContent],
) -> list[tuple[str, list[PageContent]]]:
    """Group pages into chapters of PAGES_PER_CHAPTER pages each."""
    if not pages:
        return []

    chapters: list[tuple[str, list[PageContent]]] = []
    total = len(pages)

    for start in range(0, total, PAGES_PER_CHAPTER):
        chunk = pages[start : start + PAGES_PER_CHAPTER]
        end = start + len(chunk)
        if total <= PAGES_PER_CHAPTER:
            title = "Document"
        elif start == 0:
            title = f"Pages 1–{end}"
        else:
            title = f"Pages {start + 1}–{end}"
        chapters.append((title, chunk))

    return chapters


def _markdown_inline_to_html(text: str) -> str:
    escaped = _esc(text)
    escaped = re.sub(r"`([^`]+)`", r"<code>\1</code>", escaped)
    escaped = re.sub(r"\*\*([^*]+)\*\*", r"<strong>\1</strong>", escaped)
    escaped = re.sub(r"\*([^*]+)\*", r"<em>\1</em>", escaped)
    escaped = re.sub(r"\[([^\]]+)\]\(([^)]+)\)", r'<a href="\2">\1</a>', escaped)
    return escaped


def _markdown_to_html(markdown_text: str, chapter_title: str) -> str:
    lines = markdown_text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    parts: list[str] = [f"<h1>{_esc(chapter_title)}</h1>"]
    paragraph_lines: list[str] = []
    list_items: list[str] = []
    in_code_block = False
    code_lines: list[str] = []

    def flush_paragraph() -> None:
        if paragraph_lines:
            text = " ".join(line.strip() for line in paragraph_lines).strip()
            if text:
                parts.append(f"<p>{_markdown_inline_to_html(text)}</p>")
            paragraph_lines.clear()

    def flush_list() -> None:
        if list_items:
            items_html = "".join(
                f"<li>{_markdown_inline_to_html(item)}</li>"
                for item in list_items
            )
            parts.append(f"<ul>{items_html}</ul>")
            list_items.clear()

    def flush_code_block() -> None:
        if code_lines:
            parts.append(f"<pre><code>{_esc(chr(10).join(code_lines))}</code></pre>")
            code_lines.clear()

    for raw_line in lines:
        line = raw_line.rstrip()
        stripped = line.strip()

        if stripped.startswith("```"):
            flush_paragraph()
            flush_list()
            if in_code_block:
                flush_code_block()
                in_code_block = False
            else:
                in_code_block = True
            continue

        if in_code_block:
            code_lines.append(line)
            continue

        if not stripped:
            flush_paragraph()
            flush_list()
            continue

        heading_match = re.match(r"^(#{1,3})\s+(.*)$", stripped)
        if heading_match:
            flush_paragraph()
            flush_list()
            level = len(heading_match.group(1))
            content = _markdown_inline_to_html(heading_match.group(2).strip())
            parts.append(f"<h{level}>{content}</h{level}>")
            continue

        if stripped.startswith(">"):
            flush_paragraph()
            flush_list()
            quote_text = _markdown_inline_to_html(stripped.lstrip("> ").strip())
            parts.append(f"<blockquote><p>{quote_text}</p></blockquote>")
            continue

        list_match = re.match(r"^[-*]\s+(.*)$", stripped)
        if list_match:
            flush_paragraph()
            list_items.append(list_match.group(1).strip())
            continue

        flush_list()
        paragraph_lines.append(stripped)

    flush_paragraph()
    flush_list()
    if in_code_block:
        flush_code_block()

    return "\n".join(parts)


def markdown_to_epub(markdown_bytes: bytes, filename: str = "document.md") -> bytes:
    """Convert a Markdown byte string to an EPUB3 byte string."""
    try:
        markdown_text = markdown_bytes.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ValueError("Markdown files must be UTF-8 encoded.") from exc

    if not markdown_text.strip():
        raise ValueError("The uploaded Markdown file is empty.")

    stem = Path(filename).stem
    book, css = _make_book(stem)

    chapter = epub.EpubHtml(
        title="Document",
        file_name="chapter_001.xhtml",
        lang="en",
    )
    chapter.content = (
        f'<?xml version="1.0" encoding="utf-8"?>'
        f'<!DOCTYPE html>'
        f'<html xmlns="http://www.w3.org/1999/xhtml" xml:lang="en">'
        f'<head>'
        f'<title>{_esc(stem)}</title>'
        f'<link rel="stylesheet" type="text/css" href="style/main.css"/>'
        f'</head>'
        f'<body>'
        f'{_markdown_to_html(markdown_text, "Document")}'
        f'</body>'
        f'</html>'
    ).encode("utf-8")
    chapter.add_item(css)
    book.add_item(chapter)
    book.toc = (epub.Link(chapter.file_name, chapter.title, "chapter_1"),)
    book.add_item(epub.EpubNcx())
    book.add_item(epub.EpubNav())
    book.spine = ["nav", chapter]

    buf = io.BytesIO()
    epub.write_epub(buf, book, {})
    return buf.getvalue()


def pdf_to_epub(pdf_bytes: bytes, filename: str = "document.pdf") -> bytes:
    """
    Convert a PDF byte string to an EPUB3 byte string.

    Raises:
        ValueError: if the PDF has no extractable text.
    """
    stem = Path(filename).stem

    # 1. Extract pages
    pages = _extract_pages_content(pdf_bytes)
    non_empty = [p for p in pages if not p.is_empty]

    if not non_empty:
        raise ValueError(
            "No text could be extracted from this PDF. "
            "It may be a scanned document without a text layer."
        )

    # 2. Group into chapters
    chapters = _group_into_chapters(non_empty)

    # 3. Build EPUB
    book, css = _make_book(stem)

    epub_chapters: list[epub.EpubHtml] = []

    for idx, (chapter_title, chapter_pages) in enumerate(chapters, start=1):
        html_content = _pages_to_html(chapter_pages, chapter_title)
        chapter_filename = f"chapter_{idx:03d}.xhtml"

        ep_chapter = epub.EpubHtml(
            title=chapter_title,
            file_name=chapter_filename,
            lang="en",
        )
        ep_chapter.content = (
            f'<?xml version="1.0" encoding="utf-8"?>'
            f'<!DOCTYPE html>'
            f'<html xmlns="http://www.w3.org/1999/xhtml" xml:lang="en">'
            f'<head>'
            f'<title>{_esc(chapter_title)}</title>'
            f'<link rel="stylesheet" type="text/css" href="style/main.css"/>'
            f'</head>'
            f'<body>'
            f'{html_content}'
            f'</body>'
            f'</html>'
        ).encode("utf-8")
        ep_chapter.add_item(css)
        book.add_item(ep_chapter)
        epub_chapters.append(ep_chapter)

    # 4. Navigation
    book.toc = tuple(
        epub.Link(ch.file_name, ch.title, f"chapter_{i}")
        for i, ch in enumerate(epub_chapters, start=1)
    )
    book.add_item(epub.EpubNcx())
    book.add_item(epub.EpubNav())
    book.spine = ["nav"] + epub_chapters

    # 5. Write to bytes
    buf = io.BytesIO()
    epub.write_epub(buf, book, {})
    return buf.getvalue()


# ── Excel conversion ───────────────────────────────────────────────────────────

def _cell_value_to_str(value: object) -> str:
    """Normalise a cell value to a display string."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        # Drop unnecessary decimals (e.g. 1.0 → "1")
        return str(int(value)) if value == int(value) else str(value)
    return str(value)


def _sheet_rows_openpyxl(ws) -> list[list[str]]:  # type: ignore[type-arg]
    """Extract rows from an openpyxl worksheet (values already computed)."""
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_value_to_str(cell) for cell in row])
    # Drop trailing empty rows
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    return rows


def _sheet_rows_xlrd(ws) -> list[list[str]]:  # type: ignore[type-arg]
    """Extract rows from an xlrd worksheet."""
    rows: list[list[str]] = []
    for row_idx in range(ws.nrows):
        rows.append([
            _cell_value_to_str(ws.cell(row_idx, col_idx).value)
            for col_idx in range(ws.ncols)
        ])
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    return rows


def _rows_to_html_table(rows: list[list[str]], sheet_title: str) -> str:
    """Convert sheet row data to an HTML table inside a chapter."""
    if not rows:
        return (
            f"<h1>{_esc(sheet_title)}</h1>"
            "<p class='sheet-meta'>(Empty sheet)</p>"
        )

    max_cols = max((len(r) for r in rows), default=0)
    if max_cols == 0:
        return (
            f"<h1>{_esc(sheet_title)}</h1>"
            "<p class='sheet-meta'>(Empty sheet)</p>"
        )

    def pad(row: list[str]) -> list[str]:
        return row + [""] * (max_cols - len(row))

    parts: list[str] = [
        f"<h1>{_esc(sheet_title)}</h1>",
        f"<p class='sheet-meta'>{len(rows) - 1} rows · {max_cols} columns</p>",
        '<div class="table-wrap"><table>',
        "<thead><tr>",
        *[f"<th>{_esc(c)}</th>" for c in pad(rows[0])],
        "</tr></thead><tbody>",
    ]
    for row in rows[1:]:
        parts.append("<tr>")
        parts.extend(f"<td>{_esc(c)}</td>" for c in pad(row))
        parts.append("</tr>")
    parts.append("</tbody></table></div>")
    return "\n".join(parts)


def _make_epub_chapter(
    title: str,
    filename: str,
    body_html: str,
    css: epub.EpubItem,
) -> epub.EpubHtml:
    ch = epub.EpubHtml(title=title, file_name=filename, lang="en")
    ch.content = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<!DOCTYPE html>'
        '<html xmlns="http://www.w3.org/1999/xhtml" xml:lang="en">'
        f'<head><title>{_esc(title)}</title>'
        '<link rel="stylesheet" type="text/css" href="style/main.css"/>'
        f'</head><body>{body_html}</body></html>'
    ).encode("utf-8")
    ch.add_item(css)
    return ch


def excel_to_epub(excel_bytes: bytes, filename: str = "workbook.xlsx") -> bytes:
    """
    Convert an Excel file (.xlsx or .xls) to an EPUB3 byte string.
    Each non-empty sheet becomes one chapter.

    Raises:
        ValueError: if the file has no extractable data.
    """
    stem = Path(filename).stem
    ext = Path(filename).suffix.lower()

    sheets: list[tuple[str, list[list[str]]]] = []

    if ext == ".xls":
        import xlrd  # type: ignore[import-untyped]
        wb = xlrd.open_workbook(file_contents=excel_bytes)
        for name in wb.sheet_names():
            sheets.append((name, _sheet_rows_xlrd(wb.sheet_by_name(name))))
    else:
        import openpyxl  # type: ignore[import-untyped]
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        for name in wb.sheetnames:
            sheets.append((name, _sheet_rows_openpyxl(wb[name])))
        wb.close()

    non_empty = [(name, rows) for name, rows in sheets if rows]
    if not non_empty:
        raise ValueError(
            "No data could be extracted from this Excel file. "
            "All sheets appear to be empty."
        )

    book, css = _make_book(stem)

    epub_chapters: list[epub.EpubHtml] = []
    for idx, (sheet_name, rows) in enumerate(non_empty, start=1):
        ch = _make_epub_chapter(
            title=sheet_name,
            filename=f"sheet_{idx:03d}.xhtml",
            body_html=_rows_to_html_table(rows, sheet_name),
            css=css,
        )
        book.add_item(ch)
        epub_chapters.append(ch)

    book.toc = tuple(
        epub.Link(ch.file_name, ch.title, f"sheet_{i}")
        for i, ch in enumerate(epub_chapters, start=1)
    )
    book.add_item(epub.EpubNcx())
    book.add_item(epub.EpubNav())
    book.spine = ["nav"] + epub_chapters

    buf = io.BytesIO()
    epub.write_epub(buf, book, {})
    return buf.getvalue()
